from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from pypdf import PdfReader

from src.schemas.reporting import ProfitLossReportResponse, ReportLineItem, ReportSection
from src.services.reporting.excel_export_service import ExcelExportService
from src.services.reporting.pdf_export_service import PdfExportService, format_currency_idr


def _section(code: str, name: str, amounts: list[str]) -> ReportSection:
    lines = [ReportLineItem(account_code=f"{code}{index}", line_name=f"{name} {index}", amount=Decimal(amount)) for index, amount in enumerate(amounts, start=1)]
    return ReportSection(section_code=code, section_name=name, lines=lines, subtotal=sum((line.amount for line in lines), Decimal("0.00")))


def _profit_loss() -> ProfitLossReportResponse:
    revenue = _section("41", "Pendapatan", ["123456789.12", "0.03"])
    cogs = _section("51", "HPP", ["23456789.10"])
    opex = _section("61", "Beban Operasional", ["1000000.01"])
    other = _section("71", "Pendapatan Lain", ["0.02"])
    gross = revenue.subtotal - cogs.subtotal
    operating = gross - opex.subtotal
    ebt = operating + other.subtotal
    tax = Decimal("250000.01")
    return ProfitLossReportResponse(
        organization_name="PT Presisi Nusantara", period_label="Agustus 2026",
        start_date=date(2026, 8, 1), end_date=date(2026, 8, 31), generated_at="2026-08-30",
        revenue_section=revenue, cogs_section=cogs, gross_profit=gross,
        gross_margin_percentage=Decimal("81.00"), operating_expenses_section=opex,
        operating_profit=operating, other_income_expense_section=other,
        earnings_before_tax=ebt, tax_expense=tax, net_profit=ebt - tax,
    )


def test_xlsx_formula_totals_reconcile_to_authoritative_dto_values():
    report = _profit_loss()
    workbook = load_workbook(ExcelExportService.export("profit-loss", report), data_only=False)
    assert workbook["Reconciliation"].sheet_state == "hidden"
    audit = workbook["Reconciliation"]
    authoritative = {audit.cell(row, 1).value: Decimal(str(audit.cell(row, 4).value)) for row in range(2, audit.max_row + 1)}
    assert authoritative["revenue_total"] == report.revenue_section.subtotal
    assert authoritative["gross_profit"] == report.gross_profit
    assert authoritative["net_profit"] == report.net_profit
    report_sheet = workbook["Report"]
    net_profit_row = next(row for row in range(5, report_sheet.max_row + 1) if report_sheet.cell(row, 2).value == "Laba Bersih")
    assert str(report_sheet.cell(net_profit_row, 3).value).startswith("=")
    assert workbook.calculation.fullCalcOnLoad is True


def test_pdf_totals_reconcile_exactly_to_authoritative_dto_values():
    report = _profit_loss()
    pdf = PdfExportService.export("profit-loss", report)
    text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf.read())).pages)
    assert format_currency_idr(report.revenue_section.subtotal) in text
    assert format_currency_idr(report.gross_profit) in text
    assert format_currency_idr(report.net_profit) in text
    assert "Disahkan oleh" in text
    assert "Halaman 1" in text
