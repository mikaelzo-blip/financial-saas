import io
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.schemas.reporting import (
    ProfitLossReportResponse,
    BalanceSheetReportResponse,
    TrialBalanceResponse,
    CashFlowReportResponse,
    ARAgingReportResponse,
    APAgingReportResponse
)


class ExcelExportService:
    @staticmethod
    def _apply_header_style(ws, title: str, subtitle: str, max_col: int = 4):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.cell(row=1, column=1, value=title).font = Font(name="Calibri", size=14, bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        ws.cell(row=2, column=1, value=subtitle).font = Font(name="Calibri", size=10, italic=True)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    @staticmethod
    def export_profit_loss(data: ProfitLossReportResponse) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Laba Rugi"

        ExcelExportService._apply_header_style(
            ws,
            title=f"{data.organization_name} — LAPORAN LABA RUGI",
            subtitle=f"Periode: {data.period_label}",
            max_col=3
        )

        headers = ["Kode", "Keterangan Akun / Komponen", "Jumlah (IDR)"]
        ws.append([])
        ws.append(headers)

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        for col_idx in range(1, 4):
            c = ws.cell(row=4, column=col_idx)
            c.font = header_font
            c.fill = header_fill

        row_idx = 5

        def add_section(sec_name, lines, subtotal):
            nonlocal row_idx
            ws.cell(row=row_idx, column=2, value=sec_name).font = Font(bold=True)
            row_idx += 1
            for l in lines:
                ws.cell(row=row_idx, column=1, value=l.account_code or "")
                ws.cell(row=row_idx, column=2, value=l.line_name)
                c = ws.cell(row=row_idx, column=3, value=float(l.amount))
                c.number_format = '#,##0.00'
                row_idx += 1
            # Subtotal
            ws.cell(row=row_idx, column=2, value=f"Total {sec_name}").font = Font(bold=True)
            c = ws.cell(row=row_idx, column=3, value=float(subtotal))
            c.font = Font(bold=True)
            c.number_format = '#,##0.00'
            row_idx += 2

        add_section("I. PENDAPATAN USAHA", data.revenue_section.lines, data.revenue_section.subtotal)
        add_section("II. HARGA POKOK PROYEK (HPP)", data.cogs_section.lines, data.cogs_section.subtotal)

        ws.cell(row=row_idx, column=2, value="LABA KOTOR (GROSS PROFIT)").font = Font(bold=True)
        c = ws.cell(row=row_idx, column=3, value=float(data.gross_profit))
        c.font = Font(bold=True)
        c.number_format = '#,##0.00'
        row_idx += 2

        add_section("III. BEBAN OPERASIONAL", data.operating_expenses_section.lines, data.operating_expenses_section.subtotal)

        ws.cell(row=row_idx, column=2, value="LABA USAHA (OPERATING PROFIT)").font = Font(bold=True)
        c = ws.cell(row=row_idx, column=3, value=float(data.operating_profit))
        c.font = Font(bold=True)
        c.number_format = '#,##0.00'
        row_idx += 2

        ws.cell(row=row_idx, column=2, value="LABA BERSIH TAHUN BERJALAN").font = Font(bold=True, color="047857")
        c = ws.cell(row=row_idx, column=3, value=float(data.net_profit))
        c.font = Font(bold=True, color="047857")
        c.number_format = '#,##0.00'

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 25

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def export_balance_sheet(data: BalanceSheetReportResponse) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Neraca"

        ExcelExportService._apply_header_style(
            ws,
            title=f"{data.organization_name} — LAPORAN NERACA",
            subtitle=f"Per Tanggal: {data.as_of_date}",
            max_col=3
        )

        ws.append([])
        ws.append(["Kode", "Komponen Neraca", "Jumlah (IDR)"])
        for col_idx in range(1, 4):
            c = ws.cell(row=4, column=col_idx)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

        row_idx = 5

        def add_section(sec_name, lines, subtotal):
            nonlocal row_idx
            ws.cell(row=row_idx, column=2, value=sec_name).font = Font(bold=True)
            row_idx += 1
            for l in lines:
                ws.cell(row=row_idx, column=1, value=l.account_code or "")
                ws.cell(row=row_idx, column=2, value=l.line_name)
                c = ws.cell(row=row_idx, column=3, value=float(l.amount))
                c.number_format = '#,##0.00'
                row_idx += 1
            ws.cell(row=row_idx, column=2, value=f"Total {sec_name}").font = Font(bold=True)
            c = ws.cell(row=row_idx, column=3, value=float(subtotal))
            c.font = Font(bold=True)
            c.number_format = '#,##0.00'
            row_idx += 2

        add_section("ASET LANCAR", data.current_assets.lines, data.current_assets.subtotal)
        add_section("ASET TETAP", data.fixed_assets.lines, data.fixed_assets.subtotal)

        ws.cell(row=row_idx, column=2, value="TOTAL ASET").font = Font(bold=True, color="047857")
        c = ws.cell(row=row_idx, column=3, value=float(data.total_assets))
        c.font = Font(bold=True, color="047857")
        c.number_format = '#,##0.00'
        row_idx += 2

        add_section("KEWAJIBAN JANGKA PENDEK", data.current_liabilities.lines, data.current_liabilities.subtotal)
        add_section("EKUITAS", data.equity.lines, data.equity.subtotal)

        ws.cell(row=row_idx, column=2, value="TOTAL KEWAJIBAN & EKUITAS").font = Font(bold=True, color="1E3A8A")
        c = ws.cell(row=row_idx, column=3, value=float(data.total_liabilities_and_equity))
        c.font = Font(bold=True, color="1E3A8A")
        c.number_format = '#,##0.00'

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 25

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
