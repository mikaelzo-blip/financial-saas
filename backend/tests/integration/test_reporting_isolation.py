from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from tests.reporting_support import seed_cash_profit_ledger


@pytest.mark.asyncio
async def test_report_and_export_endpoints_never_cross_tenants(client: AsyncClient, db_session: AsyncSession):
    alpha = await seed_cash_profit_ledger(db_session, "alpha-isolation", Decimal("100000000.01"))
    beta = await seed_cash_profit_ledger(db_session, "beta-isolation", Decimal("50000000.02"))
    params = {"start_date": "2026-08-01", "end_date": "2026-08-31"}

    missing_tenant = await client.get("/api/v1/reports/export/profit-loss", params={**params, "format": "xlsx"})
    assert missing_tenant.status_code == 400

    alpha_headers = {"X-Organization-ID": str(alpha.id), "Authorization": "Bearer test-session"}
    beta_headers = {"X-Organization-ID": str(beta.id), "Authorization": "Bearer test-session"}
    alpha_json = await client.get("/api/v1/reports/profit-loss", params=params, headers=alpha_headers)
    beta_json = await client.get("/api/v1/reports/profit-loss", params=params, headers=beta_headers)
    assert Decimal(str(alpha_json.json()["net_profit"])) == Decimal("100000000.01")
    assert Decimal(str(beta_json.json()["net_profit"])) == Decimal("50000000.02")

    alpha_export = await client.get("/api/v1/reports/export/profit-loss", params={**params, "format": "xlsx"}, headers=alpha_headers)
    beta_export = await client.get("/api/v1/reports/export/profit-loss", params={**params, "format": "xlsx"}, headers=beta_headers)
    assert alpha_export.status_code == beta_export.status_code == 200
    assert alpha_export.headers["content-type"].startswith("application/vnd.openxmlformats")

    def audit_value(content: bytes, key: str) -> Decimal:
        sheet = load_workbook(BytesIO(content), data_only=False)["Reconciliation"]
        row = next(row for row in range(2, sheet.max_row + 1) if sheet.cell(row, 1).value == key)
        return Decimal(str(sheet.cell(row, 4).value))

    assert audit_value(alpha_export.content, "net_profit") == Decimal("100000000.01")
    assert audit_value(beta_export.content, "net_profit") == Decimal("50000000.02")
    assert b"Beta Isolation" not in alpha_export.content
    assert b"Alpha Isolation" not in beta_export.content
